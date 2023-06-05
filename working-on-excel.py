import openpyxl as xl
from openpyxl.chart import Reference, BarChart

wb=xl.load_workbook('test.xlsx')
sheet=wb['Sheet1']

cell=sheet['a1']
cell=sheet.cell(1,1)


#print(cell.value)
#print(sheet.max_row)

for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)
    corrected=cell.value*0.800
    modified_cell=sheet.cell(row,4)
    modified_cell.value=corrected
    #print(modified_cell.value)

values=Reference(sheet,
                 min_row=2,
                 max_row=sheet.max_row,
                 min_col=4,
                 max_col=4)
chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'f2')
wb.save('test11.xlsx')